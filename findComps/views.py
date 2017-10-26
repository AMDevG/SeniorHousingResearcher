import os
import time
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from django.shortcuts import render
from .zipcodeSearch import getNearbyZips
from .models import Providerinfo
from django.http import HttpResponse

import json, requests
import simplejson

returned_data = []

def find(request):
	return render(request, 'comps/findSNFS.html' )

def getDistance(subject_address, temp_address):
	try:
		url = "https://maps.googleapis.com/maps/api/distancematrix/json?units=imperial&origins="+subject_address+"&destinations="+temp_address+"&key=AIzaSyABAGhtae7h91pf24nfyZ6eFiPThWW3W4M"
		result = requests.get(url)
		data = json.loads(result.text)
		distance = data['rows'][0]['elements'][0]['distance']['text']
		return distance
	except:
		return None





def createDownload():

	file_path = "/home/blueprintmapper/BPMapper/findComps/Comp Tables/Competitive Market2.xlsx"

	if os.path.exists(file_path):		#Creates forced download
			with open(file_path, 'rb') as fh:
				response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
				response['Content-Disposition'] = 'inline; filename=' + "Competitive Market2.xlsx"
				return response



def compsearch(request):
	counter = 0

	if request.method == 'POST':
		subject_address = request.POST.get('subj_addr', None)
		zipcode = request.POST.get('zip', None)
		radius = request.POST.get('radius', None)
		zips = getNearbyZips(zipcode, radius)  #Collects zipcodes in radius

		facility_information = search_by_zipcode(zips, subject_address) #returns dictionary of all facilities' info with matching zips
		write_excel_file(facility_information) #Creates Excel workbook containing data

		file_path = "/home/blueprintmapper/BPMapper/findComps/Comp Tables/Competitive Market2.xlsx"


		###MAY NEED TO PUT INTO OTHER FUNCTION TO REDIRECT AFTER
		return createDownload()

	return render(request, 'comps/findSNFS.html')

def search_by_zipcode(zipcodes, subject_address):
 #Accepts array of zipcodes as argument, returns dictionary of nursing home info for homes in those zipcodes
	facility_information = {}
	del returned_data[:]

	for target in zipcodes:
		base_url = "https://data.medicare.gov/resource/b27b-2uc7.json?$$app_token=Gw0QkTqNHfviiEteAF8odUR2X&provider_zip_code=" + target
		r = requests.get(base_url)
		result = json.loads(r.text)
		if len(result)!= 0:
			returned_data.append(result)

	for lst in returned_data:
		for i in range(0,len(lst)):
			data = lst[i]

			prov_name = data["provider_name"]
			facility_information[prov_name] = {}
			total_residents = float(data["number_of_residents_in_certified_beds"])
			total_beds = float(data["number_of_certified_beds"])
			reported_occ = total_residents/total_beds

			try:
				overall_rating = data["overall_rating"]
			except:
				pass

			phone_number = data["provider_phone_number"]
			sff_facility = data["special_focus_facility"]
			changed_ownership_TTM = data["provider_changed_ownership_in_last_12_months"]
			number_of_penalties = data["total_number_of_penalties"]

			facility_information[prov_name]['provider_name'] = data["provider_name"].lower().title()
			facility_information[prov_name]['provider_address'] = data["provider_address"]
			facility_information[prov_name]['provider_city'] = data["provider_city"].lower().title()
			facility_information[prov_name]['provider_state'] = data["provider_state"]
			facility_information[prov_name]['provider_zip_code'] = data["provider_zip_code"]

			formatted_address = data["provider_address"] + data["provider_city"] + data["provider_state"] + data["provider_zip_code"]
			distance = getDistance(subject_address, formatted_address)
			facility_information[prov_name]['distance'] = distance

			facility_information[prov_name]['continuing_care_retirement_community'] = data['continuing_care_retirement_community']
			facility_information[prov_name]['provider_resides_in_hospital'] = data['provider_resides_in_hospital']
			facility_information[prov_name]['CMS_rating'] = overall_rating
			facility_information[prov_name]['occupancy'] = reported_occ
			facility_information[prov_name]['total_beds'] = total_beds
			facility_information[prov_name]['phone_number'] = phone_number
			facility_information[prov_name]['sff_facility'] = sff_facility
			facility_information[prov_name]['changed_ownership_TTM'] = changed_ownership_TTM
			facility_information[prov_name]['number_of_penalties'] = number_of_penalties
	return facility_information

def write_excel_file(facility_information):

	col_counter = 2
	row_counter = 2
	target_wb = xl.load_workbook("/home/blueprintmapper/BPMapper/findComps/Comp Tables/compTemplate.xlsx")  #Need to change for server
	#target_wb = xl.load_workbook("C:\\Users\\John Berry\\Desktop\\Comp Tables\\compTemplate.xlsx")
	target_ws = target_wb["Competitive Market"]

	for key in facility_information:
		target_ws.cell(row = row_counter, column = 2).value = facility_information[key]['provider_name']
		target_ws.cell(row = row_counter, column = 3).value = facility_information[key]['provider_address']
		target_ws.cell(row = row_counter, column = 4).value = facility_information[key]['provider_city']
		target_ws.cell(row = row_counter, column = 5).value = facility_information[key]['provider_state']
		target_ws.cell(row = row_counter, column = 6).value = facility_information[key]['provider_zip_code']
		target_ws.cell(row = row_counter, column = 7).value = facility_information[key]['phone_number']
		target_ws.cell(row = row_counter, column = 8).value = "SNF"
		target_ws.cell(row = row_counter, column = 9).value = facility_information[key]['total_beds']
		target_ws.cell(row = row_counter, column = 10).value = facility_information[key]['distance']		#Distance parameter need to implement the Calculator
		target_ws.cell(row = row_counter, column = 11).value = facility_information[key]['sff_facility']
		target_ws.cell(row = row_counter, column = 12).value = facility_information[key]['occupancy']
		target_ws.cell(row = row_counter, column = 13).value = facility_information[key]['continuing_care_retirement_community']
		target_ws.cell(row = row_counter, column = 14).value = facility_information[key]['provider_resides_in_hospital']
		target_ws.cell(row = row_counter, column = 15).value = facility_information[key]['changed_ownership_TTM']
		target_ws.cell(row = row_counter, column = 16).value = facility_information[key]['number_of_penalties']
		target_ws.cell(row = row_counter, column = 17).value = facility_information[key]['CMS_rating']
		row_counter += 1

	workbook_name = ("Competitive Market2.xlsx")
	target_wb.save("/home/blueprintmapper/BPMapper/findComps/Comp Tables/"+workbook_name) ## Need to update to server
	#target_wb.save("C:\\Users\\John Berry\\Desktop\\Comp Tables\\"+workbook_name)




